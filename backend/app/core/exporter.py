"""
报告导出引擎
支持 PDF / Excel / CSV 三种格式的数据导出
"""
import os
import csv
import io
import time
import logging
from typing import Dict, Any, List, Optional, Tuple

logger = logging.getLogger(__name__)

# 导出文件暂存目录
EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports")


class ReportExporter:
    """报告导出器"""

    def __init__(self, db):
        self._db = db
        os.makedirs(EXPORT_DIR, exist_ok=True)

    async def export_csv(self, owner: str, repo: str, collection: str,
                         query: Dict = None, fields: List[str] = None) -> str:
        """导出 CSV 文件"""
        if self._db is None or self._db.db is None:
            raise RuntimeError("数据库未连接")

        self._cleanup_old_exports()

        query = query or {}
        if owner:
            query["owner"] = owner
        if repo:
            query["repo"] = repo

        docs = []
        cursor = self._db.db[collection].find(query, {"_id": 0})
        async for doc in cursor:
            docs.append(self._flatten_doc(doc))

        if not docs:
            raise ValueError("没有可导出的数据")

        # 如果未指定字段，使用第一个文档的所有键
        if not fields:
            fields = list(docs[0].keys())

        filename = f"{owner}_{repo}_{collection}_{int(time.time())}.csv"
        filepath = os.path.join(EXPORT_DIR, filename)

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for doc in docs:
                # 将所有值转为字符串，处理嵌套对象
                row = {k: str(v) if v is not None else "" for k, v in doc.items() if k in fields}
                writer.writerow(row)

        logger.info(f"CSV 导出完成: {filename}, {len(docs)} 条记录")
        return filepath

    async def export_excel(self, owner: str, repo: str,
                           sheets_config: List[Dict]) -> str:
        """导出 Excel（多 Sheet）"""
        from openpyxl import Workbook

        if self._db is None or self._db.db is None:
            raise RuntimeError("数据库未连接")

        self._cleanup_old_exports()

        wb = Workbook()
        first_sheet = True

        for sheet_cfg in sheets_config:
            collection = sheet_cfg["collection"]
            query = dict(sheet_cfg.get("query", {}))
            if owner:
                query["owner"] = owner
            if repo:
                query["repo"] = repo
            sheet_fields = sheet_cfg.get("fields")
            sheet_name = sheet_cfg.get("name", collection)[:31]  # Excel Sheet 名称限制 31 字符

            docs = []
            cursor = self._db.db[collection].find(query, {"_id": 0})
            async for doc in cursor:
                docs.append(self._flatten_doc(doc))

            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            if not docs:
                ws.append(["无数据"])
                continue

            fields = sheet_fields or list(docs[0].keys())
            ws.append(fields)

            for doc in docs:
                row = [str(doc.get(f, "")) if doc.get(f) is not None else "" for f in fields]
                ws.append(row)

            # 自动调整列宽
            for col_idx, field in enumerate(fields, 1):
                max_len = len(str(field))
                for doc in docs[:100]:
                    val = str(doc.get(field, ""))
                    max_len = max(max_len, min(len(val), 50))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        filename = f"{owner}_{repo}_data_{int(time.time())}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        wb.save(filepath)

        logger.info(f"Excel 导出完成: {filename}")
        return filepath

    async def export_pdf(self, owner: str, repo: str, report_type: str,
                         report_data: Dict, date_range: Tuple[str, str] = None) -> str:
        """导出 PDF 报告"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        self._cleanup_old_exports()

        filename = f"{owner}_{repo}_{report_type}_{int(time.time())}.pdf"
        filepath = os.path.join(EXPORT_DIR, filename)

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()

        # 尝试注册中文字体
        try:
            # macOS 系统字体
            font_path = "/System/Library/Fonts/PingFang.ttc"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("PingFang", font_path))
                cn_font = "PingFang"
            else:
                cn_font = "Helvetica"
        except Exception:
            cn_font = "Helvetica"

        # 自定义样式
        title_style = ParagraphStyle(
            "CNTitle", parent=styles["Title"],
            fontName=cn_font, fontSize=18, spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            "CNHeading", parent=styles["Heading2"],
            fontName=cn_font, fontSize=14, spaceAfter=8, textColor=HexColor("#1890ff"),
        )
        body_style = ParagraphStyle(
            "CNBody", parent=styles["Normal"],
            fontName=cn_font, fontSize=10, spaceAfter=6,
        )

        story = []

        # 标题
        report_titles = {
            "cicd": "CI/CD 洞察报告",
            "review_quality": "Review 质量报告",
            "project_health": "项目健康度报告",
            "trend_alerts": "趋势预警报告",
            "all": "综合分析报告",
        }
        title = report_titles.get(report_type, "分析报告")
        story.append(Paragraph(f"{title} — {owner}/{repo}", title_style))

        date_info = ""
        if date_range and date_range[0] and date_range[1]:
            date_info = f"统计周期: {date_range[0]} ~ {date_range[1]}"
        story.append(Paragraph(date_info or f"生成时间: {report_data.get('generated_at', '')}", body_style))
        story.append(Spacer(1, 12))

        # 根据报告类型渲染不同内容
        if report_type == "project_health":
            self._build_pdf_health(story, report_data, heading_style, body_style, cn_font)
        elif report_type == "review_quality":
            self._build_pdf_review(story, report_data, heading_style, body_style, cn_font)
        elif report_type == "trend_alerts":
            self._build_pdf_alerts(story, report_data, heading_style, body_style, cn_font)
        else:
            # 通用渲染：遍历所有字段
            self._build_pdf_generic(story, report_data, heading_style, body_style, cn_font)

        doc.build(story)
        logger.info(f"PDF 导出完成: {filename}")
        return filepath

    def _build_pdf_generic(self, story, data, heading_style, body_style, cn_font):
        """通用 PDF 内容构建"""
        for key, value in data.items():
            if key in ("error", "generated_at"):
                continue
            story.append(Paragraph(str(key).replace("_", " ").title(), heading_style))
            if isinstance(value, dict):
                self._add_table(story, value, cn_font)
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                self._add_list_table(story, value, cn_font)
            else:
                story.append(Paragraph(str(value), body_style))
            story.append(Spacer(1, 8))

    def _build_pdf_health(self, story, data, heading_style, body_style, cn_font):
        """项目健康度 PDF 内容"""
        story.append(Paragraph(f"总体评分: {data.get('overall_score', 'N/A')} 分 ({data.get('overall_grade', 'N/A')})", body_style))
        story.append(Spacer(1, 8))

        for dim in data.get("dimensions", []):
            story.append(Paragraph(f"{dim.get('name', '')}: {dim.get('score', 0)} 分 ({dim.get('grade', '')})", body_style))
            if dim.get("suggestion"):
                story.append(Paragraph(f"  建议: {dim['suggestion']}", body_style))
        story.append(Spacer(1, 8))

        insights = data.get("insights", [])
        if insights:
            story.append(Paragraph("洞察与建议", heading_style))
            for ins in insights:
                story.append(Paragraph(f"- {ins}", body_style))

    def _build_pdf_review(self, story, data, heading_style, body_style, cn_font):
        """Review 质量 PDF 内容"""
        metrics = data.get("summary", {})
        story.append(Paragraph("概览指标", heading_style))
        self._add_table(story, metrics, cn_font)
        story.append(Spacer(1, 8))

        reviewers = data.get("top_reviewers", [])
        if reviewers:
            story.append(Paragraph("Top Reviewers", heading_style))
            self._add_list_table(story, reviewers[:10], cn_font)

    def _build_pdf_alerts(self, story, data, heading_style, body_style, cn_font):
        """趋势预警 PDF 内容"""
        alerts = data.get("alerts", [])
        story.append(Paragraph(f"预警数量: {len(alerts)}", body_style))
        story.append(Spacer(1, 8))

        for alert in alerts:
            severity = alert.get("severity", "info")
            color = "#f5222d" if severity == "critical" else "#faad14" if severity == "warning" else "#1890ff"
            story.append(Paragraph(f"[{severity.upper()}] {alert.get('title', '')}", heading_style))
            story.append(Paragraph(alert.get("description", ""), body_style))
            if alert.get("suggestion"):
                story.append(Paragraph(f"建议: {alert['suggestion']}", body_style))
            story.append(Spacer(1, 6))

    def _add_table(self, story, data: Dict, cn_font: str):
        """将字典转为 PDF 表格"""
        if not data:
            return
        table_data = [[str(k), str(v)] for k, v in data.items() if k != "_id"]
        t = Table(table_data, colWidths=[6 * cm, 10 * cm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), cn_font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (0, -1), HexColor("#f5f5f5")),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e8e8e8")),
        ]))
        story.append(t)

    def _add_list_table(self, story, data: List[Dict], cn_font: str):
        """将列表转为 PDF 表格"""
        if not data:
            return
        # 取所有键的并集（限制前 8 列）
        all_keys = list(dict.fromkeys(k for d in data for k in d.keys() if k != "_id"))[:8]
        table_data = [all_keys]
        for item in data[:50]:
            table_data.append([str(item.get(k, "")) for k in all_keys])

        col_width = min(4 * cm, 16 * cm / len(all_keys))
        t = Table(table_data, colWidths=[col_width] * len(all_keys))
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), cn_font),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1890ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e8e8e8")),
        ]))
        story.append(t)

    def _flatten_doc(self, doc: Dict, prefix: str = "") -> Dict:
        """将嵌套文档展平为一级字典"""
        items = {}
        for k, v in doc.items():
            new_key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                items.update(self._flatten_doc(v, new_key))
            elif isinstance(v, list):
                items[new_key] = str(v)[:200]
            else:
                items[new_key] = v
        return items

    def _cleanup_old_exports(self, max_age_hours: int = 24):
        """清理过期的导出文件"""
        try:
            if not os.path.exists(EXPORT_DIR):
                return
            now = time.time()
            for f in os.listdir(EXPORT_DIR):
                filepath = os.path.join(EXPORT_DIR, f)
                if os.path.isfile(filepath) and (now - os.path.getmtime(filepath)) > max_age_hours * 3600:
                    os.remove(filepath)
        except Exception as e:
            logger.warning(f"清理导出文件失败: {e}")

    # ==================== 工作流仿真报告导出 ====================

    def export_simulation_pdf(self, summary: Dict[str, Any], step_results: List[Dict[str, Any]]) -> str:
        """导出工作流仿真 PDF 报告"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        self._cleanup_old_exports()

        plugin_name = summary.get("plugin_name", "unknown")
        sim_id = summary.get("simulation_id", "unknown")
        filename = f"simulation_{plugin_name}_{sim_id}_{int(time.time())}.pdf"
        filepath = os.path.join(EXPORT_DIR, filename)

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()

        # 中文字体
        try:
            font_path = "/System/Library/Fonts/PingFang.ttc"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont("PingFang", font_path))
                cn_font = "PingFang"
            else:
                cn_font = "Helvetica"
        except Exception:
            cn_font = "Helvetica"

        title_style = ParagraphStyle("SimTitle", parent=styles["Title"], fontName=cn_font, fontSize=18, spaceAfter=12)
        heading_style = ParagraphStyle("SimH2", parent=styles["Heading2"], fontName=cn_font, fontSize=14, spaceAfter=8, textColor=HexColor("#1890ff"))
        body_style = ParagraphStyle("SimBody", parent=styles["Normal"], fontName=cn_font, fontSize=10, spaceAfter=6)
        small_style = ParagraphStyle("SimSmall", parent=styles["Normal"], fontName=cn_font, fontSize=8, spaceAfter=4, textColor=HexColor("#666666"))

        story = []

        # (a) 报告头部
        story.append(Paragraph(f"工作流仿真报告 — {plugin_name}", title_style))
        meta_text = (
            f"角色: {summary.get('persona', 'N/A')} | "
            f"仿真 ID: {sim_id} | "
            f"时间: {summary.get('compared_at', 'N/A')}"
        )
        story.append(Paragraph(meta_text, body_style))
        story.append(Spacer(1, 12))

        # (b) 统计概览
        story.append(Paragraph("统计概览", heading_style))
        total_skills = sum(
            len(s.get("skills_used", [])) + len(s.get("skills_missing", []))
            for s in step_results
        )
        used_skills = sum(len(s.get("skills_used", [])) for s in step_results)
        coverage = f"{round(used_skills / total_skills * 100)}%" if total_skills > 0 else "N/A"

        stats_data = [
            ["总通过率", f"{round(summary.get('overall_pass_rate', 0) * 100)}%"],
            ["断点总数", str(summary.get("total_breakpoints", 0))],
            ["CRITICAL 断点", str(summary.get("critical_breakpoints", 0))],
            ["Skill 覆盖率", coverage],
            ["匹配反模式", str(len(summary.get("antipatterns_matched", [])))],
            ["Token 消耗", str(summary.get("total_tokens", 0))],
            ["预估成本", f"${summary.get('estimated_cost_usd', 0)}"],
            ["步骤数", str(len(step_results))],
        ]
        stats_table = Table(stats_data, colWidths=[6 * cm, 10 * cm])
        stats_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), cn_font),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BACKGROUND", (0, 0), (0, -1), HexColor("#f0f5ff")),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e8e8e8")),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 16))

        # (c) 断点列表
        all_breakpoints = []
        for s in step_results:
            for bp in s.get("breakpoints", []):
                all_breakpoints.append({**bp, "step_name": s.get("step_name", "")})

        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        all_breakpoints.sort(key=lambda b: severity_order.get(b.get("severity", "LOW"), 9))

        if all_breakpoints:
            story.append(Paragraph("断点列表", heading_style))
            bp_header = ["步骤", "严重性", "类别", "描述", "建议"]
            bp_rows = [bp_header]
            for bp in all_breakpoints[:50]:
                desc = (bp.get("description", "") or "")[:80]
                rec = (bp.get("recommendation", "") or "")[:60]
                bp_rows.append([
                    bp.get("step_name", ""),
                    bp.get("severity", ""),
                    bp.get("category", ""),
                    desc,
                    rec,
                ])
            bp_table = Table(bp_rows, colWidths=[2.5 * cm, 2 * cm, 3 * cm, 5 * cm, 3.5 * cm])
            bp_style = [
                ("FONTNAME", (0, 0), (-1, -1), cn_font),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1890ff")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e8e8e8")),
            ]
            for idx, bp in enumerate(all_breakpoints[:50], 1):
                sev = bp.get("severity", "")
                if sev == "CRITICAL":
                    bp_style.append(("BACKGROUND", (0, idx), (-1, idx), HexColor("#fff1f0")))
                elif sev == "HIGH":
                    bp_style.append(("BACKGROUND", (0, idx), (-1, idx), HexColor("#fff7e6")))
            bp_table.setStyle(TableStyle(bp_style))
            story.append(bp_table)
            story.append(Spacer(1, 16))

        # (d) 分步详情
        story.append(Paragraph("分步详情", heading_style))
        for s in step_results:
            pass_pct = round(s.get("simulated_pass_rate", 0) * 100)
            story.append(Paragraph(
                f"{s.get('step_name', '')} (通过率: {pass_pct}%)",
                ParagraphStyle("StepH", parent=styles["Heading3"], fontName=cn_font, fontSize=11, textColor=HexColor("#333333")),
            ))

            if s.get("skills_used"):
                story.append(Paragraph(f"Skills Used: {', '.join(s['skills_used'])}", small_style))
            if s.get("skills_missing"):
                story.append(Paragraph(f"Skills Missing: {', '.join(s['skills_missing'])}", small_style))

            for bp in s.get("breakpoints", []):
                sev_color = "#f5222d" if bp.get("severity") == "CRITICAL" else "#fa8c16" if bp.get("severity") == "HIGH" else "#faad14"
                story.append(Paragraph(
                    f"<font color='{sev_color}'>[{bp.get('severity', '')}]</font> "
                    f"<font color='#999999'>[{bp.get('category', '')}]</font> "
                    f"{bp.get('description', '')}",
                    small_style,
                ))

            llm_summary = (s.get("llm_response_summary", "") or "")[:200]
            if llm_summary:
                story.append(Paragraph(f"LLM: {llm_summary}", small_style))
            story.append(Spacer(1, 8))

        # (e) Skill 热力图
        heatmap = summary.get("skill_heatmap", {})
        if heatmap:
            story.append(PageBreak())
            story.append(Paragraph("Skill 利用热力图", heading_style))
            # 收集所有 skill 名称
            all_skills = sorted(set(sk for step_heat in heatmap.values() for sk in step_heat.keys()))
            step_names = []
            for s in step_results:
                sid = s.get("step_id", "")
                if sid in heatmap:
                    step_names.append(s.get("step_name", sid))

            hm_header = ["Skill"] + step_names
            hm_rows = [hm_header]
            for skill in all_skills:
                row = [skill]
                for s in step_results:
                    sid = s.get("step_id", "")
                    val = heatmap.get(sid, {}).get(skill, 0)
                    row.append(f"{round(val * 100)}%" if val > 0 else "-")
                hm_rows.append(row)

            col_w = min(3 * cm, 16 * cm / max(len(hm_header), 1))
            hm_table = Table(hm_rows, colWidths=[4 * cm] + [col_w] * len(step_names))
            hm_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), cn_font),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1890ff")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e8e8e8")),
            ]))
            story.append(hm_table)
            story.append(Spacer(1, 16))

        # (f) 匹配反模式
        antipatterns = summary.get("antipatterns_matched", [])
        if antipatterns:
            story.append(Paragraph("匹配的反模式", heading_style))
            for ap in antipatterns:
                sev = ap.get("severity", "")
                sev_color = "#f5222d" if sev == "CRITICAL" else "#fa8c16" if sev == "HIGH" else "#faad14"
                suscept = round(ap.get("susceptibility", 0) * 100)
                story.append(Paragraph(
                    f"<font color='{sev_color}'>[{sev}]</font> {ap.get('name', '')} — 易感性: {suscept}%",
                    body_style,
                ))
                if ap.get("mitigation"):
                    story.append(Paragraph(f"建议: {ap['mitigation']}", small_style))
                story.append(Spacer(1, 4))

        doc.build(story)
        logger.info(f"仿真 PDF 导出完成: {filename}")
        return filepath

    def export_simulation_markdown(self, summary: Dict[str, Any], step_results: List[Dict[str, Any]]) -> str:
        """导出工作流仿真 Markdown 报告"""
        self._cleanup_old_exports()

        plugin_name = summary.get("plugin_name", "unknown")
        sim_id = summary.get("simulation_id", "unknown")

        lines = []
        lines.append(f"# 工作流仿真报告 — {plugin_name}\n")

        # 基本信息
        lines.append("## 基本信息\n")
        lines.append("| 字段 | 值 |")
        lines.append("|------|-----|")
        lines.append(f"| 插件 | {plugin_name} |")
        lines.append(f"| 角色 | {summary.get('persona', 'N/A')} |")
        lines.append(f"| 仿真 ID | {sim_id} |")
        lines.append(f"| 时间 | {summary.get('compared_at', 'N/A')} |")
        lines.append("")

        # 统计概览
        total_skills = sum(
            len(s.get("skills_used", [])) + len(s.get("skills_missing", []))
            for s in step_results
        )
        used_skills = sum(len(s.get("skills_used", [])) for s in step_results)
        coverage = f"{round(used_skills / total_skills * 100)}%" if total_skills > 0 else "N/A"

        lines.append("## 统计概览\n")
        lines.append("| 指标 | 值 |")
        lines.append("|------|-----|")
        lines.append(f"| 总通过率 | {round(summary.get('overall_pass_rate', 0) * 100)}% |")
        lines.append(f"| 断点总数 | {summary.get('total_breakpoints', 0)} |")
        lines.append(f"| CRITICAL 断点 | {summary.get('critical_breakpoints', 0)} |")
        lines.append(f"| Skill 覆盖率 | {coverage} |")
        lines.append(f"| 匹配反模式 | {len(summary.get('antipatterns_matched', []))} |")
        lines.append(f"| Token 消耗 | {summary.get('total_tokens', 0):,} |")
        lines.append(f"| 预估成本 | ${summary.get('estimated_cost_usd', 0)} |")
        lines.append("")

        # 断点列表
        all_breakpoints = []
        for s in step_results:
            for bp in s.get("breakpoints", []):
                all_breakpoints.append({**bp, "step_name": s.get("step_name", "")})

        severity_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        all_breakpoints.sort(key=lambda b: severity_order.get(b.get("severity", "LOW"), 9))

        if all_breakpoints:
            lines.append("## 断点列表\n")
            lines.append("| 步骤 | 严重性 | 类别 | 描述 | 建议 |")
            lines.append("|------|--------|------|------|------|")
            for bp in all_breakpoints:
                desc = bp.get("description", "").replace("|", "\\|")[:80]
                rec = bp.get("recommendation", "").replace("|", "\\|")[:60]
                lines.append(f"| {bp.get('step_name', '')} | **{bp.get('severity', '')}** | {bp.get('category', '')} | {desc} | {rec} |")
            lines.append("")

        # 步骤详情
        lines.append("## 步骤详情\n")
        for s in step_results:
            pass_pct = round(s.get("simulated_pass_rate", 0) * 100)
            lines.append(f"### {s.get('step_name', '')} (通过率: {pass_pct}%)\n")

            if s.get("skills_used"):
                lines.append(f"- **Skills Used**: {', '.join(s['skills_used'])}")
            if s.get("skills_missing"):
                lines.append(f"- **Skills Missing**: {', '.join(s['skills_missing'])}")

            for bp in s.get("breakpoints", []):
                lines.append(f"- [{bp.get('severity', '')}] [{bp.get('category', '')}] {bp.get('description', '')}")
                if bp.get("recommendation"):
                    lines.append(f"  - 建议: {bp['recommendation']}")

            llm_summary = (s.get("llm_response_summary", "") or "")[:300]
            if llm_summary:
                lines.append(f"\n> LLM 回复: {llm_summary}\n")
            lines.append("")

        # Skill 热力图
        heatmap = summary.get("skill_heatmap", {})
        if heatmap:
            lines.append("## Skill 利用热力图\n")
            all_skills = sorted(set(sk for step_heat in heatmap.values() for sk in step_heat.keys()))
            step_header = ["Skill"] + [s.get("step_name", s.get("step_id", "")) for s in step_results if s.get("step_id", "") in heatmap]
            lines.append("| " + " | ".join(step_header) + " |")
            lines.append("| " + " | ".join(["------"] * len(step_header)) + " |")
            for skill in all_skills:
                vals = [skill]
                for s in step_results:
                    sid = s.get("step_id", "")
                    val = heatmap.get(sid, {}).get(skill, 0)
                    vals.append(f"{round(val * 100)}%" if val > 0 else "-")
                lines.append("| " + " | ".join(vals) + " |")
            lines.append("")

        # 匹配反模式
        antipatterns = summary.get("antipatterns_matched", [])
        if antipatterns:
            lines.append("## 匹配的反模式\n")
            for ap in antipatterns:
                suscept = round(ap.get("susceptibility", 0) * 100)
                lines.append(f"### {ap.get('name', '')} [{ap.get('severity', '')}]\n")
                lines.append(f"- 易感性: {suscept}%")
                if ap.get("mitigation"):
                    lines.append(f"- 建议: {ap['mitigation']}")
                lines.append("")

        # 写入文件
        filename = f"simulation_{plugin_name}_{sim_id}_{int(time.time())}.md"
        filepath = os.path.join(EXPORT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        logger.info(f"仿真 Markdown 导出完成: {filename}")
        return filepath
